import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import warnings
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

chrome_options = Options()
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"]) # 셀레니움 로그 무시
warnings.filterwarnings("ignore", category=DeprecationWarning) # Deprecated warning 무시 

browser = webdriver.Chrome("c:/chromedriver.exe", options=chrome_options)

# 웹사이트 열기
# browser.get('https://new.land.naver.com/rooms?ms=37.538617,127.082375,14&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT')
browser.get('https://new.land.naver.com/rooms?ms=37.546892,127.103025,16&a=APT:OPST:ABYG:OBYG:GM:OR:VL:DDDGG:JWJT:SGJT:HOJT&e=RETAIL&aa=SMALLSPCRENT')
print(browser.title)
print(browser.current_url)

# 로딩이 끝날때까지 10초정도 기다린다.
browser.implicitly_wait(10) 

# 페이지 다운 입력
print("python")
scrollTarget = browser.find_element(By.CLASS_NAME, 'item_link')
print(scrollTarget)
scrollTarget.click()

print("JS")
scrollBox = browser.execute_script("return document.querySelector('.item_list').scrollTop") # .item_list로 스크롤 위치 확인가능

print(scrollBox)

print("키 적용")

# 무한 스크롤

#스크롤 전 높이
beforeScroll = 0

while True:
    scrollTarget.send_keys(Keys.END)
    time.sleep(1)
    #스크롤 후 높이
    afterScroll = browser.execute_script("return document.querySelector('.item_list').scrollTop") 

    if afterScroll == beforeScroll:
        break
    beforeScroll = afterScroll
itemList = browser.find_elements(By.CLASS_NAME, 'item_link')
number = 0

종류 = browser.execute_script('return document.querySelector(".item_inner > .item_link > .item_title").innerText')
가격 = browser.execute_script('return document.querySelector(".item_inner > .item_link > .price_line").innerText')
info1 = browser.execute_script('return document.querySelectorAll(".item_inner > .item_link > .info_area > .line")[0].innerText')
info2 = browser.execute_script('return document.querySelectorAll(".item_inner > .item_link > .info_area > .line")[1].innerText')
tag = browser.execute_script('return document.querySelector(".item_inner > .item_link > .tag_area").innerText')
agentInfo1 = browser.execute_script('return document.querySelectorAll(".item_inner > .cp_area > .cp_area_inner > .agent_info")[0].innerText')
agentInfo2 = browser.execute_script('return document.querySelectorAll(".item_inner > .cp_area > .cp_area_inner > .agent_info")[1].innerText')


소재지 = browser.execute_script('return document.querySelectorAll(".table_td")[0].innerText')
매물특징 = browser.execute_script('return document.querySelectorAll(".table_td")[1].innerText')
공급_전용면적 = browser.execute_script('return document.querySelectorAll(".table_td")[2].innerText')
해당층_총층 = browser.execute_script('return document.querySelectorAll(".table_td")[3].innerText')
방수_욕실수 = browser.execute_script('return document.querySelectorAll(".table_td")[4].innerText')
월관리비 = browser.execute_script('return document.querySelectorAll(".table_td")[5].innerText')
관리비포함 = browser.execute_script('return document.querySelectorAll(".table_td")[6].innerText')
입주가능일 = browser.execute_script('return document.querySelectorAll(".table_td")[7].innerText')
융자금 = browser.execute_script('return document.querySelectorAll(".table_td")[8].innerText')
사용승인일 = browser.execute_script('return document.querySelectorAll(".table_td")[9].innerText')
방향 = browser.execute_script('return document.querySelectorAll(".table_td")[10].innerText')
주차가능여부 = browser.execute_script('return document.querySelectorAll(".table_td")[11].innerText')
방구조 = browser.execute_script('return document.querySelectorAll(".table_td")[12].innerText')
복층여부 = browser.execute_script('return document.querySelectorAll(".table_td")[13].innerText')
건축물용도 = browser.execute_script('return document.querySelectorAll(".table_td")[14].innerText')
매물번호 = browser.execute_script('return document.querySelectorAll(".table_td")[15].innerText')
총주차대수 = browser.execute_script('return document.querySelectorAll(".table_td")[16].innerText')
중개사_이름 = browser.execute_script('return document.querySelector(".info_agent_title .info_title").innerText')
중개사_사진 = browser.execute_script('return document.querySelector(".info_agent_photo .info_photo_image").src')
중개사_직함 = browser.execute_script('return document.querySelectorAll(".info_agent_wrap > .info_agent")[0].firstElementChild.innerText')
중개사_직원명 = browser.execute_script('return document.querySelectorAll(".info_agent_wrap > .info_agent")[0].firstElementChild.nextElementSibling.innerText')
중개사_주소 = browser.execute_script('return document.querySelectorAll(".info_agent_wrap > .info_agent")[0].firstElementChild.nextElementSibling.nextElementSibling.lastElementChild.innerText')
중개사_전화 = browser.execute_script('return document.querySelectorAll(".info_agent_wrap > .info_agent")[1].firstElementChild.nextElementSibling.innerText')


# print(info1) # 필요 없음
# print(info2) # 필요없음
# print(tag) # 필요없음
# print(agentInfo1) # 필요없음
# print(agentInfo2) # 필요없음

print(종류) # 일반원룸
print(가격) # 월세 500/30
print(소재지) # 주소
print(매물특징) # 설명
print(공급_전용면적) #
print(해당층_총층)
print(방수_욕실수)
print(월관리비)
print(관리비포함)
print(입주가능일)
print(융자금)
print(사용승인일)
print(방향)
print(주차가능여부)
print(방구조)
print(복층여부)
print(중개사_이름)
print(중개사_사진)
print(중개사_직함)
print(중개사_직원명)
print(중개사_주소)
print(중개사_전화)

# 엑셀파일 만들기
wb = openpyxl.Workbook()

# 엑셀 워크시트 만들기
ws = wb.create_sheet('광진구')

# 데이터 추가하기
ws['A1'] = '종류'
ws['B1'] = '가격'
ws['C1'] = '공급/전용면적'
ws['D1'] = '층/총층'
ws['E1'] = '방수/욕실수'
ws['F1'] = '월 관리비'
ws['G1'] = '관리비포함'
ws['H1'] = '입주 가능일'
ws['I1'] = '융자금'
ws['J1'] = '사용 승인일'
ws['K1'] = '방향'
ws['L1'] = '주차가능여부'
ws['M1'] = '방구조'
ws['N1'] = '복층여부'
ws['O1'] = '중개사_이름'
ws['P1'] = '중개사_사진'
ws['Q1'] = '중개사_직함'
ws['R1'] = '중개사_직원명'
ws['S1'] = '중개사_주소'
ws['T1'] = '중개사_전화'

# 20개
abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T'] 

# 20개
budongsan = [종류, 가격, 공급_전용면적, 해당층_총층, 방수_욕실수, 월관리비, 관리비포함, 입주가능일, 융자금, 사용승인일, 방향, 주차가능여부, 방구조, 복층여부, 중개사_이름, 중개사_사진, 중개사_직함, 중개사_직원명, 중개사_주소, 중개사_전화]
for i in range(5):
    for j in range(20):
        ws[abc[j] + str(i+2)] = budongsan[j]

# for i in range(19):
#     for j in abc:
#         print(j + str(i+2))
#         ws[j + str(i+2)] = budongsan[j]


# 엑셀 저장하기
wb.save('부동산_data.xlsx')

# !한개만 먼저 해보고 성공하면 반복문 돌려서 전부해
# for item in itemList:
#     number = number + 1
#     # item.click()
#     # 찾은 item을 클릭해서 정보창을 열고,
#     # 정보창의 데이터를 추출
#     print(number)
#     print(item)

# 아이템 클릭
# itemList = browser.find_elements(By.CLASS_NAME, 'item_link')
# for element in itemList: 
#     print(element);
    # element.click()






